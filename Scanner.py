# Analyse d' relevé de compte bancaire au format PDF
# 1) Conversion pdf -> xml
# par  .\env\Scripts\pdf2txt.py -o releve.xml releve.pdf

import re
import sys
import os
import xml.etree.cElementTree as ET
import shutil
from operator import itemgetter

from Scanner_Engine import Scanner
from pathlib import Path
from os import path
from glob import glob

from openpyxl import load_workbook

from logger import logger

import argparse
import importlib

import ScannerTools
from config import Config
from ScannerTools import *


def export_to_excel (config, params, scanner):
    """
        Exporte les données dans un fichier Excel
    """
    # Export dans le fichier excel
    excel_filename = os.path.join(config.output_dir,"Releves-Bancaires.xlsx")
    exists = os.path.isfile(excel_filename)   
    if not exists:
        logger.error(f"Le fichier XLS : {excel_filename} est absent, il est impossible de rajouter les données d'export au sein de celui ci")
        exit(99)
        
    
    # sélectionne la bonne feuille
    workbook = load_workbook(excel_filename)


    sheet_name = params.releve["config"]["excel_sheet"]
    
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        logger.error(f"Impossible de trouver la bonne feuille de données {sheet_name} au sein du fichier XLS : {excel_filename} ")
        exit(99)
            

    last_rows = sheet.max_row+1
    for line in scanner.output_lines:
        elems = line.split(";")
        for col, value in enumerate(elems, start=1):
            
            if col==3:
                value = int(value)
            elif col in (8,9):               
              value = strFloatToExport(value)
                  

            
            sheet.cell(row=last_rows,column = col,value=value)
        last_rows=last_rows+1
    workbook.save(excel_filename)



def parse_XML_document(config, params, xml_in):

    try:
        # Cas où certains libellés ne sont pas affichés sur la même ligne exactement
        # on gère e tolérance
        line_tolerance = params.releve["config"]["offset_tolerance_ligne"]

        # Récupération des éléments de la page
        tree = ET.ElementTree(file=xml_in)
        pages = tree.getroot()

        if pages.tag != "pages":
            sys.exit("ERROR: pages.tag is %s instead of pages!" % pages.tag)

        page_index = 0

        # Création du scanner de document
        scanner = Scanner(config, params, xml_in)

        for index, page in enumerate(pages):
            # On récupère tous les éléments "textline"
            textlines = page.findall("./textbox/textline")

            # Parcours à travers les "textlines"
            lines = []
            for textline in textlines:
                # Obtention des limites des "textline"
                line_bods = [float(s) for s in textline.attrib["bbox"].split(",")]

                # Récupération des caractères formant la chaîne
                chars = list(textline)

                # Combine tous les caractères en e seule chaine
                line_text = ""
                for char in chars:
                    line_text = str(line_text) + str(char.text)

                # Supprime les multiples espaces
                line_text = re.sub(" +", " ", line_text.strip())

                # Description de la ligne
                line = {
                    "left": line_bods[0],
                    "top": line_bods[1],
                    "text": line_text,
                }
                lines.append(line)

            # Effectue  premier tri des lignes par x et par y (Attention en inverse car en pdf 0 est en bas)
            # sort the lines by left, then reverse top position
            lines.sort(key=itemgetter("left"))
            lines.sort(key=itemgetter("top"), reverse=True)

            # Certaines données sur la meme ligne n'ont pas en réalité la même ordonnée y
            # on gère cela en modifiant l'ordonnée suivant e tolérance
            line_top = lines[0]["top"]
            for line in lines:
                if abs(line["top"] - line_top) < line_tolerance:
                    line["top"] = line_top
                else:
                    line_top = line["top"]

            # On effectue  tri après les modification de lignes éventuelles
            lines.sort(key=itemgetter("left"))
            lines.sort(key=itemgetter("top"), reverse=True)

            # === Export debug de la page ===
            if config.logs:
                debug_filename = "DBG-" + Path(xml_in).stem
                page_index += 1
                page_filename =  os.path.join(config.output_dir,"_tmp",debug_filename)+ "-"+ str(page_index).zfill(4)+ ".txt"
                logger.debug(f"Export description xml de la page {page_filename}")
                with open(page_filename,"w",encoding="utf-8",) as f:
                    for item in lines:
                        f.write("%s\n" % item)

            # On demande au scanner d'analyser la page
            logger.info(f"Debut d'analyse de la page {index}")
            if not scanner.process_page(page_index, lines):
                logger.error(f"Impossible de traiter la page {index} du document")
                exit(99)
        
        export_filename = os.path.join(config.output_dir,"Releves.csv")
        exists = os.path.isfile(export_filename)        
        with open(export_filename, "a") as file:
            if not exists:
                file.write(scanner.line_export.get_header()+"\n")
            for line in scanner.output_lines:
                file.write(str(line) + "\n")
                
                
        # export_to_excel(config, params, scanner)    
        


    finally:
        # releve_id = scanner.releve_id
        logger.info("=== Fin de traitement ===")


def file_process(config):
    # importe le bon module de paramétrage en fonction du scanner défini
    logger.info(f"Utilisation du module {config.scanner}")
    params = importlib.import_module(config.scanner)

    # supprime le fichier temporaire de traitement xml
    output_file = os.path.join(config.output_dir,"_tmp", f"releve-{config.scanner}.xml")
    logger.debug(f"output file {output_file}")
    if os.path.isfile(output_file):
        # supprime l'ancien document xml
        os.remove(output_file)

    # Lance la génération du fichier xml correpondant à l'extract PDF
    # attention pdf2txt est dans l'environnement python utilisé

    pdf2txt_file = os.path.join(Path(sys.executable).parent, "pdf2txt.py")
    if not os.path.isfile(pdf2txt_file):
        logger.error(f"Le script {pdf2txt_file} est absent, vérifiez que vous exécutez le script dans son environnement propre et que vous avez installé les modules requis!")
        exit(99)

    os.system(f'{os.path.join(Path(sys.executable).parent,"pdf2txt.py")} -o "{output_file}" "{config.file_pdf}"')

    # traitement
    releve_id = parse_XML_document(config, params, output_file)


    logger.info("=============")
    logger.info(f"Ancien Solde : {params.dynamic.get('ancien_solde',0)}")
    logger.info(f"Total crédit : {params.dynamic.get('total_credit',0)}")
    logger.info(f"Total débit : {params.dynamic.get('total_debit',0)}")


    

    
                    


    # Nettoyage du fichier xml
    if os.path.isfile(output_file):
        os.remove(output_file)

    # on copie le relevés en le renommant dans répertoire de sortie traité
    # shutil.copy(
    #     config.file_pdf, os.path.join("releves/Traites/", str(releve_id).zfill(3) + ".pdf")
    # )

    # On déplace le fichier original dans ToDelte
    # shutil.move(filename, os.path.join("releves/ToDelete/"))


# Fonction de tri personnalisée pour trier par nom de fichier, puis par date


def custom_sort(file_path):
    file_name = path.basename(file_path)
    return file_name.lower(), os.path.getmtime(file_path)


def directory_process(config):
    pdf_files = glob(path.join(config.directory, "*.{}".format("pdf")))
    nb_files = len(pdf_files)
    logger.info(f"Nombre de fichiers PDF à traiter {nb_files}")

    # Trier les fichiers en utilisant la fonction de tri personnalisée
    sorted_files = sorted(pdf_files, key=custom_sort)

    file_on_error = []
    i = 0
    for file in sorted_files:
        i += 1
        logger.info(f"Traitement du fichier {file} {i}/{nb_files}")
        try:
            config.file_pdf = file
            file_process(config)
        except:
            print("<!> Erreur sur fichier", file)
            file_on_error.append(file)

    print("=============================================")
    print("   FIN DE TRAITEMENT ")
    print("=============================================")
    if len(file_on_error) > 0:
        print("")
        print("=== Fichiers en erreur ===")
        for f in file_on_error:
            print(f)


if __name__ == "__main__":
    parser = argparse.ArgumentParser("Scanner de relevés bancaires")

    # Création d' groupe de paramètres mutuellement exclusifs
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("-f", "--file-pdf", help="fichier de relevé PDF à traiter")
    group.add_argument("-d", "--directory", help="Répertoire des relevés PDF à traiter")    
    parser.add_argument("-o", "--output-dir", help="Nom du répertoire de sortie", required=True)

    parser.add_argument("-s", "--scanner", help="Scanner à utiliser [LCL|BoursoBank]")
    parser.add_argument("-r", "--reference-compte", help="référence compte", required=True)

    parser.add_argument("--logs", help="Génère les fichier de debug", action="store_true")
        
    args = parser.parse_args()
    

    config = Config(
        # répertoire du script
        os.path.dirname(os.path.abspath(__file__)),       
        # fichier pdf à importer
        args.file_pdf,
        # répertoire pour les imports de fichiers
        args.directory,
        # Répertoire de sortie
        args.output_dir,
        # scanner à utiliser
        args.scanner,
        # référence compte
        args.reference_compte, 
        # True si log
        args.logs,
    )

    
    logger.info(f"Fichier : {config.file_pdf}")
    logger.info(f"Scanner : {config.scanner}")

    ScannerTools.create_output_structure(config.output_dir)


    if args.file_pdf:
        logger.info(f"Traitement du fichier {config.file_pdf}")
        file_process(config)
        logger.info(f"Import fichier {config.file_pdf}")
    if args.directory:
        logger.info(f"--- Traitement du répertoire {config.directory}")
        directory_process(config)
